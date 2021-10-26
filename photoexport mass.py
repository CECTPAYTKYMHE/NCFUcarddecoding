import openpyxl
#from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader
import os
cwd = os.path.abspath('D:\\GIT\\Files\\work')
files = os.listdir(cwd)

for file in files:
    try:
        filein = cwd + '\\' + file
        wb = openpyxl.load_workbook(filein)
        ws = wb['ФИОиФотоСотр']
        i = 2 #начальная строка
        row = ws.max_row
        image_loader = SheetImageLoader(ws)
        ws.cell(row=1, column=5).value = 'Фотография №'
        while i != row + 1: #проверить конечную строку!!!!
            ID = ws.cell(row=i, column=1)
            ID = str(ID.value)
            ID = ID.replace('-','')
            ID = ID.upper()
            image = image_loader.get('E' + str(i)) #клетка с фотографией
            image.save('d:\\git\\files\\photo\\employee\\' + ID + '.jpg') #экспорт фото
            ws.cell(row=i, column=5).value = ID +'.jpg' #имяфото.jpg
            ws.cell(row=i, column=1).value = ID #удаление '-' в NCFUGUID
            i += 1
        ws._images = [] # удаление всех фото из файла
        wb.save(filename = filein)
        print(f'{file} успех')
    except:
        print(f'{file} завершился неудачно')

