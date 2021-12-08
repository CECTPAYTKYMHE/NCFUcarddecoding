import openpyxl
from openpyxl_image_loader import SheetImageLoader
import os
import csv
import io
import glob
from PIL import Image
cwd = 'D:\\GIT\\Files\\student\\spo\\'
files = [f for f in glob.glob(cwd + '**/*.xlsx', recursive=True)]
Image.MAX_IMAGE_PIXELS = None
try:
    logs = open(f'D:\\GIT\\Files\\student\\errorlogs.csv', 'x')
    logs.write('Ошибка фотографии в файле, Строка, ФИО\n')
    logs.close()
except:
    pass
for file in files:
    try:
        #filein = cwd + '\\' + file
        wb = openpyxl.load_workbook(file)
        ws = wb['ФИОиФото']
        i = 3 #начальная строка
        row = ws.max_row
        SheetImageLoader._images.clear()
        image_loader = SheetImageLoader(ws)
        ws.cell(row=1, column=1).value = 'ID'
        ws.cell(row=1, column=2).value = 'Фамилия'
        ws.cell(row=1, column=3).value = 'Имя'
        ws.cell(row=1, column=4).value = 'Отчество'
        ws.cell(row=1, column=5).value = 'Фотография №'
        #ws.cell(row=1, column=6).value = 'Role'
        while i != row + 1: #проверить конечную строку!!!!
            ID = ws.cell(row=i, column=1)
            ID = str(ID.value)
            ID = ID.replace('-','')
            ID = ID.upper()
            try:
                ws.unmerge_cells(start_row=i, start_column=5, end_row=i, end_column=6)
            except:
                pass
            try:
                image = image_loader.get('E' + str(i)) #клетка с фотографией
                if image.format.lower() in ['jpg', 'jpeg']:
                    image.save('d:\\git\\files\\student\\spo\\photo\\' + ID + '.jpg') #экспорт фото
                    ws.cell(row=i, column=5).value = ID + '.jpg' #имяфото.jpg
                else:
                    logs = open(f'D:\\GIT\\Files\\student\\errorlogs.csv','a')
                    print(f'<Неправильная фотография в файле {file} в строке {i-1} у студента {ws.cell(row=i, column=2).value} {ws.cell(row=i, column=3).value} {ws.cell(row=i, column=4).value}>')
                    logs.write(f'{file}, {i-1}, {ws.cell(row=i, column=2).value} {ws.cell(row=i, column=3).value} {ws.cell(row=i, column=4).value} \n')
                    logs.close()
                    pass
            except:
                logs = open(f'D:\\GIT\\Files\\student\\errorlogs.csv','a')
                print(f'<Неправильная фотография в файле {file} в строке {i-1} у студента {ws.cell(row=i, column=2).value} {ws.cell(row=i, column=3).value} {ws.cell(row=i, column=4).value}>')
                logs.write(f'{file}, {i-1}, {ws.cell(row=i, column=2).value} {ws.cell(row=i, column=3).value} {ws.cell(row=i, column=4).value} \n')
                logs.close()
                pass
            #ws.cell(row=i, column=6).value = 'Student'
            ws.cell(row=i, column=1).value = ID #удаление '-' в NCFUGUID
            ws.cell(row=i-1, column=1).value = ws.cell(row=i, column=1).value
            ws.cell(row=i-1, column=2).value = ws.cell(row=i, column=2).value
            ws.cell(row=i-1, column=3).value = ws.cell(row=i, column=3).value
            ws.cell(row=i-1, column=4).value = ws.cell(row=i, column=4).value
            ws.cell(row=i-1, column=5).value = ws.cell(row=i, column=5).value
            i += 1
        ws._images = [] # удаление всех фото из файла
        #ws.delete_cols(7, 2) # удаление столбцов с 7 по 9(7 + 2)
        wb.save(filename = file)
        wb.close()
        print(f'{file} успех')
    except:
        print(f'{file} в строке {i} завершился неудачно')


    
    
    
    

