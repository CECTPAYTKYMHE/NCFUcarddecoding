import openpyxl
from openpyxl_image_loader import SheetImageLoader
import os,glob
cwd = 'D:\\GIT\\Files\\employee180222\\'
files = [f for f in glob.glob(cwd + '**/*.xlsx', recursive=True)]

for file in files:
    try:
        # filein = cwd + '\\' + file
        wb = openpyxl.load_workbook(file)
        ws = wb.worksheets[0]
        i = 2 #начальная строка
        row = ws.max_row
        image_loader = SheetImageLoader(ws)
        ws.cell(row=1, column=5).value = 'Фотография №'
        ws.cell(row=1, column=6).value = 'Role'
        ws.cell(row=1, column=7).value = ''
        while i != row + 1: #проверить конечную строку!!!!
            ws.cell(row=i, column=6).value = 'Employee'
            ID = ws.cell(row=i, column=1)
            ID = str(ID.value)
            ID = ID.replace('-','')
            ID = ID.upper()
            try:
                image = image_loader.get('H' + str(i)) #клетка с фотографией
                if image.format.lower() in ['jpg', 'jpeg']:
                    image.save(f'{cwd}photo\\{ID}.jpg') #экспорт фото
                    ws.cell(row=i, column=5).value = ID + '.jpg' #имяфото.jpg
                else:
                    logs = open(f'{cwd}errorlogs.csv','a')
                    print(f'<Неправильная фотография в файле {file} в строке {i} у сотрудника {ws.cell(row=i, column=2).value} {ws.cell(row=i, column=3).value} {ws.cell(row=i, column=4).value}>')
                    logs.write(f'{file}, {i}, {ws.cell(row=i, column=2).value} {ws.cell(row=i, column=3).value} {ws.cell(row=i, column=4).value} \n')
                    logs.close()
                    pass
            except:
                logs = open(f'{cwd}errorlogs.csv','a')
                print(f'<Неправильная фотография в файле {file} в строке {i} у сотрудника {ws.cell(row=i, column=2).value} {ws.cell(row=i, column=3).value} {ws.cell(row=i, column=4).value}>')
                logs.write(f'{file}, {i}, {ws.cell(row=i, column=2).value} {ws.cell(row=i, column=3).value} {ws.cell(row=i, column=4).value} \n')
                logs.close()
                pass
            # ws.cell(row=i, column=5).value = ID +'.jpg' #имяфото.jpg
            ws.cell(row=i, column=1).value = ID #удаление '-' в NCFUGUID
            i += 1
        ws._images = [] # удаление всех фото из файла
        ws.delete_cols(7, 2)
        wb.save(filename = file)
        print(f'{file} успех')
    except:
        print(f'{file} {i} завершился неудачно')