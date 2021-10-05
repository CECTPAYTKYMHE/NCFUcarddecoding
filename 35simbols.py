import openpyxl
from openpyxl.cell import cell

#filein = input('Полный путь до файла > ') 
filein = 'D:\\GIT\\Files\\cpamexport\\employ.xlsx'
wb = openpyxl.load_workbook(filein)
ws = wb['Лист1']

i = 2 #начальная строка
row = ws.max_row
n = 0
while i != row:
    cell_obj = ws.cell(row=i, column=5)
    if len(cell_obj.value) > 35:
        lenght = len(cell_obj.value)
        cell_obj = cell_obj.value.split()
        n = len(cell_obj) / 2
        n = int(n)
        part1 = cell_obj[0:n + 1]
        part2 = cell_obj[n + 1:len(cell_obj)]
        part1 = ' '.join(part1)
        part2 = ' '.join(part2)
        if len(part1) > 35 or len(part2) > 35:
            print(f'"{part1}" или "{part2}" в строке {i} больше 35 символов')
            break
        else:
            ws.cell(row=i, column=5).value = part1
            ws.cell(row=i, column=6).value = part2
    i+=1
wb.save(filename = filein)

    

        
