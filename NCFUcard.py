import openpyxl

def decoding(ID):
    #ID = "{0:x}".format(ID) #перевод в hex
    ID = list(ID) #делаем из строки список
    i = 6 #начинаем с 6 индекса в списке
    IDnew = []
    while i > -1: #цикл переворачивания байт
        IDnew.append(ID[i:i+2])
        i -= 2
    ncfuid = list(map(''.join, IDnew)) #обьединение списка списков
    ncfuid = ''.join(ncfuid) #из списка в строку
    ncfuid = int(ncfuid, 16)#из hex в dec
    return ncfuid
filein = input('Полный путь до файла > ')    
wb = openpyxl.load_workbook(filein)
ws = wb['Лист1']

i = 1
row = ws.max_row
try:
    while i != row + 1: #перебор значений 4 столбца, перекодировка и запись обратно в книгу
        cell_obj = ws.cell(row=i, column=4)
        cell_obj = str(cell_obj.value)
        #cell_obj = int(cell_obj)
        ws.cell(row=i, column=4).value = decoding(cell_obj)
        i += 1
    wb.save(filename = filein)
except:
    print('Что-то пошло не так!')
    
    
    
    


