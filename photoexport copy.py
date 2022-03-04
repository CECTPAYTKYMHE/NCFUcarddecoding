import openpyxl
from openpyxl.styles import PatternFill

#filein = input('Полный путь до файла > ') 
print('lol')
filein = 'D:\\GIT\\f_929621c7cda1ba0a.xlsx'
filein2 = 'D:\\GIT\\f_616621c7cd9be25a.xlsx'
wb = openpyxl.load_workbook(filein)
ws = wb.worksheets[0]
wb2 = openpyxl.load_workbook(filein2)
ws2 = wb2.worksheets[0]
tele = {}
i = 4 #начальная строка
row2 = 30
row1 = 30
k = 1

while i != row1 + 1:
    name = f'{ws2.cell(row=i, column=2).value} {ws2.cell(row=i, column=3).value}'
    if ws2.cell(row=i, column=4).value != None:
        name += f' {ws2.cell(row=i, column=4).value}'
    phone = ws2.cell(row=i, column=10).value
    tele[name] = phone
    i += 1
# # ws.insert_cols(2)
color = openpyxl.styles.colors.Color(rgb='FFEAF6FF')
while k != row2 + 1: #проверить конечную строку!!!!
    ID = ws.cell(row=k, column=1)
    # print(ID.value)
    if (ID.fill.start_color.index) == 'FFEAF6FF':
        # print('yes')
        try:
            print(ID.value)
            print(tele[ID.value])
            ws.cell(row=k, column=5).value = tele[ID.value]
        except:
            pass
    k += 1
wb.save(filename = filein)

print(tele)


