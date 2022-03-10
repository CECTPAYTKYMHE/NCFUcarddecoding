import openpyxl
import pandas as pd
import glob
cwd = 'D:/GIT/Files/student010322/last/'
files = [f for f in glob.glob(cwd + '**/*.xlsx', recursive=True)]
id = 0
for file in files:
    id += 1
    wb = openpyxl.load_workbook(file)
    ws = wb.worksheets[0]
    ws.insert_rows(1)
    i = 2
    row = ws.max_row + 1
    ws.cell(row=1, column=1).value = 'Personnel ID'
    ws.cell(row=1, column=2).value = 'Last Name'
    ws.cell(row=1, column=3).value = 'First Name'
    ws.cell(row=1, column=4).value = 'Middle Name'
    ws.cell(row=1, column=7).value = 'Photo'
    ws.cell(row=1, column=8).value = 'Personnel Type'
    ws.cell(row=1, column=9).value = 'Card #'
    ws.cell(row=1, column=10).value = 'Access Level'
    ws.cell(row=1, column=11).value = 'Badge Format'
    ws.cell(row=1, column=12).value = 'Badge Type'
    ws.cell(row=1, column=13).value = 'Watch level'
    try:
        while i != row:
            photo = ws.cell(row=i, column=7).value
            try:
                cardid = ws.cell(row=i, column=9).value
                ws.cell(row=i, column=9).value = int(cardid)
            except:
                log = open(f'{cwd}logs.txt', 'a')
                log.write(f'Отсутствует номер карты в файле {file} в строке {i-1}\n')
                log.close()
            ws.cell(row=i, column=7).value = f'{cwd}photo\\' + photo
            ws.cell(row=i, column=8).value = 'Student'
            ws.cell(row=i, column=10).value = 'All turnstile'
            ws.cell(row=i, column=11).value = '34bit_noFAC'
            ws.cell(row=i, column=12).value = 'Standard'
            ws.cell(row=i, column=13).value = 'Студенты'
            i += 1
        wb.save(filename = f'{cwd}готовое\\{id}.xlsx')
        filexslx = f'{cwd}готовое\\{id}.xlsx'
        data_xls = pd.read_excel(filexslx, dtype=str, index_col=None)
        data_xls.to_csv(f'{cwd}готовое\\csv\\{id}.csv', encoding='cp1251', index=False)
    except:
        print(f'Какаято проблема в {file} в строке {i}')
        pass



