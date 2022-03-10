import openpyxl
import pandas as pd

filein = 'D:/GIT/Files/employee160222/ФИОиФотоСотр_16_02_22 (1).xlsx'   
wb = openpyxl.load_workbook(filein)
ws = wb.worksheets[0]
i = 2
row = ws.max_row + 1
ws.cell(row=1, column=6).value = 'Personnel Type'
ws.cell(row=1, column=7).value = 'Access Level'
ws.cell(row=1, column=8).value = 'Badge Number'
ws.cell(row=1, column=9).value = 'Badge Format'
ws.cell(row=1, column=10).value = 'Badge Type'
ws.cell(row=1, column=11).value = 'Watch level'
while i != row:
    photo = ws.cell(row=i, column=5).value
    ws.cell(row=i, column=5).value = 'D:/GIT/Files/employee160222/photo/' + photo
    ws.cell(row=i, column=6).value = 'Employee - Full Time'
    ws.cell(row=i, column=7).value = 'All turnstile'
    ws.cell(row=i, column=9).value = '34bit_noFAC'
    ws.cell(row=i, column=10).value = 'Standard'
    ws.cell(row=i, column=11).value = 'Сотрудники'
    i += 1


wb.save(filename = filein)
data_xls = pd.read_excel(filein, dtype=str, index_col=None)
data_xls.to_csv('D:/GIT/Files/employee160222/employee0303.csv', encoding='cp1251', index=False)



