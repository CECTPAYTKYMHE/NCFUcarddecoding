import openpyxl
from openpyxl_image_loader import SheetImageLoader
import pandas as pd
import os
cwd = os.path.abspath('D:\\GIT\\Files\\work')
files = os.listdir(cwd)

for file in files:
    filein = cwd + '\\' + file
    #filein = 'D:\\GIT\\Files\\cpamexport\\student.xlsx'
    #filein = input('Полный путь до файла > ')    
    wb = openpyxl.load_workbook(filein)
    ws = wb['Лист1']

    i = 2
    row = ws.max_row
    image_loader = SheetImageLoader(ws)
    try:
        ws.cell(row=1, column=1).value = 'ID'
        ws.cell(row=1, column=6).value = 'Должность'
        ws.cell(row=1, column=8).value = 'Фотография №'
        while i != row + 1:
            #print(i)
            ID = ws.cell(row=i, column=1)
            ID = str(ID.value)
            ID = ID.replace('-','')
            ID = ID.upper()
            image = image_loader.get('H' + str(i)) #клетка с фотографией
            image.save('d:\\git\\files\\photo2\\' + ID + '.jpg') #экспорт фото
            ws.cell(row=i, column=8).value = ID +'.jpg' #имяфото.jpg
            ws.cell(row=i, column=1).value = ID #удаление '-' в NCFUGUID
            ws._images = [] #удаляем все фотки из xlsx
            cell_obj = ws.cell(row=i, column=5) #дробление должности на несколько клеток если кол-во символов больше 35
            if len(cell_obj.value) > 35:
                lenght = len(cell_obj.value)
                cell_obj = cell_obj.value.split()
                n = len(cell_obj) / 2
                n = int(n)
                #if len(cell_obj[n-1]) <= 3:
                    #n += 1
                part1 = cell_obj[0:n]
                part2 = cell_obj[n:len(cell_obj)]
                
                part1 = ' '.join(part1)
                part2 = ' '.join(part2)
                if len(part1) > 35:
                        print(f'"{part1}" в файле {file} в строке {i} больше 35 символов')
                        #break
                ws.cell(row=i, column=5).value = part1
                ws.cell(row=i, column=6).value = part2
                cell_obj = ws.cell(row=i, column=6)
                if len(cell_obj.value) > 35:
                    lenght = len(cell_obj.value)
                    cell_obj = cell_obj.value.split()
                    n = len(cell_obj) / 2
                    n = int(n)
                    if len(cell_obj[n-1]) <= 3:
                     n += 1
                    part1 = cell_obj[0:n]
                    part2 = cell_obj[n:len(cell_obj)]
                    part1 = ' '.join(part1)
                    part2 = ' '.join(part2)
                    if len(part1) > 35 or len(part2) > 35:
                        print(f'"{part1}" или "{part2}" в файле {file} в строке {i} больше 35 символов')
                        #break
                    ws.cell(row=i, column=6).value = part1
                    ws.cell(row=i, column=7).value = part2
                    cell_obj5 = ws.cell(row=i, column=5)
                    cell_obj6 = ws.cell(row=i, column=6)
                    cell_obj7 = ws.cell(row=i, column=7) 
                    if len(cell_obj5.value) > 35 or len(cell_obj6.value) > 35 or len(cell_obj7.value) > 35:
                        print("Где-то все таки больше")
            i += 1       
        wb.save(filename = filein)
        #data_xls = pd.read_excel(filein, dtype=str, index_col=None)
        #data_xls.to_csv('d:\\git\\files\\' + file + '.csv', encoding='cp1251', index=False)
    except:
        print('Что-то пошло не так!')