import openpyxl
import pandas as pd
import glob
cwd = 'D:\\GIT\\card\\tocpam'
files = [f for f in glob.glob(cwd + '**/*.xlsx', recursive=False)]
id = 0
for file in files:
    id += 1
    wb = openpyxl.load_workbook(file)
    ws = wb.worksheets[0]
    i = 2
    row = ws.max_row + 1
    while i != row:
        if ws.cell(row=i, column=6).value == 'Student':
            ws.cell(row=i, column=13).value = 'Студенты'
        elif ws.cell(row=i, column=6).value == 'Employee - Full Time':
            ws.cell(row=i, column=13).value = 'Сотрудники'
        else:
            print(f'hz kto {file}, {i}')
        i += 1
    wb.save(filename = f'D:\\GIT\\card\\tocpam\\готовое\\{id}.xlsx')
    filexslx = f'D:\\GIT\\card\\tocpam\\готовое\\{id}.xlsx'
    data_xls = pd.read_excel(filexslx, dtype=str, index_col=None)
    data_xls.to_csv(f'D:\\GIT\\card\\tocpam\\готовое\\csv\\{id}.csv', encoding='cp1251', index=False)