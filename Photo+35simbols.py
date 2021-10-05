import openpyxl
from openpyxl_image_loader import SheetImageLoader
import pandas as pd

filein = 'D:\\GIT\\Files\\cpamexport\\employee.xlsx'
#filein = input('Полный путь до файла > ')    
wb = openpyxl.load_workbook(filein)
ws = wb['Лист1']

i = 2
row = ws.max_row
image_loader = SheetImageLoader(ws)
try:
    ws.cell(row=1, column=1).value = 'ID'
    ws.cell(row=1, column=6).value = 'Должность'
    ws.cell(row=1, column=8).value = 'Фотография №'
    while i != row:
        ID = ws.cell(row=i, column=1)
        ID = str(ID.value)
        ID = ID.replace('-','')
        ID = ID.upper()
        image = image_loader.get('F' + str(i)) #клетка с фотографией
        image.save('d:\\git\\files\\photo\\' + ID + '.jpg') #экспорт фото
        ws.cell(row=i, column=8).value = 'd:\\git\\files\\photo\\' + ID +'.jpg' #имяфото.jpg
        ws.cell(row=i, column=1).value = ID #удаление '-' в NCFUGUID
        ws._images = []
        cell_obj = ws.cell(row=i, column=5)
        if len(cell_obj.value) > 35:
            lenght = len(cell_obj.value)
            cell_obj = cell_obj.value.split()
            n = len(cell_obj) / 2
            n = int(n)
            part1 = cell_obj[0:n + 1]
            part2 = cell_obj[n + 1:len(cell_obj)]
            part1 = ' '.join(part1)
            part2 = ' '.join(part2)
            if len(part1) > 35 or len(part2) > 35:
                print(f'"{part1}" или "{part2}" больше 35 символов')
                break
            else:
                ws.cell(row=i, column=5).value = part1
                ws.cell(row=i, column=6).value = part2
        i += 1
    
    wb.save(filename = filein)
    data_xls = pd.read_excel(filein, dtype=str, index_col=None)
    data_xls.to_csv('d:\\git\\files\\export.csv', encoding='utf-8', index=False)

except:
    print('Что-то пошло не так!')