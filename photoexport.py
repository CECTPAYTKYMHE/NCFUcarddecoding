import openpyxl
from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader

#filein = input('Полный путь до файла > ') 
filein = 'D:\\GIT\\Files\\cpamexport\\student.xlsx'
wb = openpyxl.load_workbook(filein)
ws = wb['Лист1']


i = 2 #начальная строка
row = ws.max_row
image_loader = SheetImageLoader(ws)
while i != row + 1: #проверить конечную строку!!!!
    ID = ws.cell(row=i, column=1)
    ID = str(ID.value)
    ID = ID.replace('-','')
    image = image_loader.get('E' + str(i)) #клетка с фотографией
    image.save('d:\\git\\files\\photo\\' + ID + '.jpg') #экспорт фото
    ws.cell(row=i, column=5).value = ID +'.jpg' #имяфото.jpg
    ws.cell(row=i, column=1).value = ID #удаление '-' в NCFUGUID
    i += 1
ws._images = [] # удаление всех фото из файла
wb.save(filename = filein)

